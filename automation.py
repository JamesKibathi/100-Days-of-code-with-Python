import openpyxl

inv_file = openpyxl.load_workbook("Inventory.xlsx")

product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
products_under10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    # calculation number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("Adding a new supplier")
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic - products with inventory less that 10
    if inventory < 10:
        products_under10_inv[int(product_num)] = int(inventory)
        print(products_under10_inv)



